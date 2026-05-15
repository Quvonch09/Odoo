import base64
from io import BytesIO

from openpyxl import load_workbook

from odoo import _, fields, models
from odoo.exceptions import ValidationError


class StudentExcelImportWizard(models.TransientModel):
    _name = 'student.excel.import.wizard'
    _description = 'Student Excel Import Wizard'

    file = fields.Binary(string='Excel fayl', required=True)
    file_name = fields.Char(string='Fayl nomi')

    HEADER_ALIASES = {
        'name': 'Ism-familiya',
        'age_years': 'Yosh',
        'phone': 'Telefon raqam',
        'parent_name': 'Ota-onasining ismi',
        'parent_phone': 'Ota-onasining telefon raqami',
        'batch_id': 'Guruh',
    }

    def action_import_file(self):
        self.ensure_one()
        if not self.file:
            raise ValidationError(_("Excel fayl yuklanmagan."))

        if self.file_name and not self.file_name.lower().endswith('.xlsx'):
            raise ValidationError(_("Faqat .xlsx formatdagi fayl yuklang."))

        workbook = load_workbook(
            filename=BytesIO(base64.b64decode(self.file)),
            data_only=True,
        )
        worksheet = workbook.active
        headers = [self._normalize_cell(cell.value) for cell in worksheet[1]]
        expected = list(self.HEADER_ALIASES.values())

        if headers[:len(expected)] != expected:
            raise ValidationError(_(
                "Excel sarlavhalari noto'g'ri.\nTo'g'ri tartib: %s"
            ) % ", ".join(expected))

        student_model = self.env['op.student'].sudo()
        batch_model = self.env['op.batch'].sudo()
        imported_count = 0

        for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            row_values = [self._normalize_cell(value) for value in row[:len(expected)]]
            if not any(row_values):
                continue

            full_name, age_value, phone, parent_name, parent_phone, batch_name = row_values
            if not full_name:
                raise ValidationError(_("%s-qator: Ism-familiya bo'sh bo'lishi mumkin emas.") % row_number)
            if not phone:
                raise ValidationError(_("%s-qator: Telefon raqam bo'sh bo'lishi mumkin emas.") % row_number)

            batch = False
            if batch_name:
                batch = batch_model.search([('name', '=', batch_name)], limit=1)
                if not batch:
                    raise ValidationError(_("%s-qator: '%s' guruhi topilmadi.") % (row_number, batch_name))

            first_name, last_name = self._split_name(full_name)
            vals = {
                'name': full_name,
                'first_name': first_name,
                'last_name': last_name,
                'phone': phone,
                'parent_name': parent_name,
                'parent_phone': parent_phone,
            }
            if age_value not in (None, ''):
                try:
                    vals['age_years'] = int(float(age_value))
                except (TypeError, ValueError):
                    raise ValidationError(_("%s-qator: Yosh son bo'lishi kerak.") % row_number)
            if batch:
                vals['batch_id'] = batch.id

            student = student_model.search([('phone', '=', phone)], limit=1)
            if student:
                student.write(vals)
            else:
                student_model.create(vals)
            imported_count += 1

        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Import tugadi'),
                'message': _('%s ta student muvaffaqiyatli yuklandi.') % imported_count,
                'type': 'success',
                'sticky': False,
            }
        }

    def _normalize_cell(self, value):
        if value is None:
            return ''
        if isinstance(value, str):
            return value.strip()
        return value

    def _split_name(self, full_name):
        parts = [part for part in (full_name or '').split() if part]
        if not parts:
            return '', ''
        if len(parts) == 1:
            return parts[0], ''
        return parts[0], ' '.join(parts[1:])
